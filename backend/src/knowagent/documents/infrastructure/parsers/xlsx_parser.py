from __future__ import annotations

from collections.abc import Iterable, Sequence
from datetime import date, datetime, time
from io import BytesIO
from typing import Protocol
from uuid import UUID

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]

from knowagent.documents.domain.models import (
    ParsedBlock,
    ParsedDocument,
    SourceLocator,
    SourceType,
)
from knowagent.documents.errors import DocumentParseError, ParseErrorCode
from knowagent.documents.infrastructure.parsers.base import (
    DocumentContext,
    ParserLimits,
    column_letter,
    make_parsed_document,
    make_table_row_block,
    normalized_extension,
    normalized_media_type,
    validate_office_archive,
)


class _Worksheet(Protocol):  # pylint: disable=too-few-public-methods
    title: str
    max_row: int
    max_column: int

    def iter_rows(self, *, values_only: bool) -> Iterable[Sequence[object]]: ...


class _Workbook(Protocol):  # pylint: disable=too-few-public-methods
    worksheets: Sequence[_Worksheet]

    def close(self) -> None: ...


class XlsxDocumentParser:
    source_type = SourceType.XLSX

    def __init__(self, limits: ParserLimits | None = None) -> None:
        self._limits = limits or ParserLimits()

    def supports(self, *, media_type: str, filename: str) -> bool:
        expected = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return (
            normalized_extension(filename) == ".xlsx"
            and normalized_media_type(media_type) == expected
        )

    def parse(
        self,
        *,
        content: bytes,
        document_id: UUID,
        document_version_id: UUID,
    ) -> ParsedDocument:
        validate_office_archive(content, self._limits)
        context = DocumentContext(document_id, document_version_id)
        workbook: _Workbook | None = None
        try:
            workbook = load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
            if len(workbook.worksheets) > self._limits.max_xlsx_sheets:
                raise DocumentParseError(
                    ParseErrorCode.RESOURCE_LIMIT_EXCEEDED,
                    "Excel 工作表数量超过解析上限",
                )
            blocks = self._extract_blocks(workbook, context)
        except (InvalidFileException, KeyError, ValueError, OSError, SyntaxError) as exc:
            raise DocumentParseError(ParseErrorCode.INVALID_FILE, "Excel 文件结构无效") from exc
        finally:
            if workbook is not None:
                workbook.close()

        if not blocks:
            raise DocumentParseError(ParseErrorCode.EMPTY_DOCUMENT, "Excel 没有可索引内容")
        return make_parsed_document(
            context=context,
            source_type=self.source_type,
            blocks=blocks,
            parser_name="openpyxl",
            distribution="openpyxl",
        )

    def _extract_blocks(  # pylint: disable=too-many-locals
        self,
        workbook: _Workbook,
        context: DocumentContext,
    ) -> list[ParsedBlock]:
        blocks: list[ParsedBlock] = []
        estimated_cells = 0
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            max_row = int(worksheet.max_row)
            max_column = int(worksheet.max_column)
            if max_row > self._limits.max_xlsx_rows_per_sheet:
                raise DocumentParseError(
                    ParseErrorCode.RESOURCE_LIMIT_EXCEEDED,
                    f"工作表 {worksheet.title} 的行数超过解析上限",
                )
            if max_column > self._limits.max_xlsx_columns:
                raise DocumentParseError(
                    ParseErrorCode.RESOURCE_LIMIT_EXCEEDED,
                    f"工作表 {worksheet.title} 的列数超过解析上限",
                )
            estimated_cells += max_row * max_column
            if estimated_cells > self._limits.max_xlsx_cells:
                raise DocumentParseError(
                    ParseErrorCode.RESOURCE_LIMIT_EXCEEDED,
                    "Excel 单元格数量超过解析上限",
                )

            first_content_row: int | None = None
            for row_index, values in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1,
            ):
                serialized = tuple(_serialize_cell(value) for value in values)
                populated = [index for index, value in enumerate(serialized, start=1) if value]
                if not populated:
                    continue
                if first_content_row is None:
                    first_content_row = row_index
                first_column = populated[0]
                last_column = populated[-1]
                visible_values = serialized[first_column - 1 : last_column]
                block_index = len(blocks)
                cell_range = (
                    f"{column_letter(first_column)}{row_index}:"
                    f"{column_letter(last_column)}{row_index}"
                )
                locator = SourceLocator(
                    document_id=context.document_id,
                    document_version_id=context.document_version_id,
                    source_type=SourceType.XLSX,
                    block_index=block_index,
                    sheet_name=str(worksheet.title),
                    cell_range=cell_range,
                    table_index=sheet_index,
                    table_row_start=row_index,
                    table_row_end=row_index,
                )
                blocks.append(
                    make_table_row_block(
                        block_index=block_index,
                        source_type=SourceType.XLSX,
                        text=" | ".join(visible_values),
                        locator=locator,
                        table_id=sheet_index,
                        row_index=row_index,
                        is_header=row_index == first_content_row,
                    )
                )
        return blocks


def _serialize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).replace("\u00a0", " ").split())
